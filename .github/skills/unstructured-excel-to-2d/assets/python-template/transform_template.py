import csv
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parents[4]
INPUT_FILE = BASE_DIR / "input.xlsx"
OUTPUT_XLSX = BASE_DIR / "output_cleaned.xlsx"
OUTPUT_CSV = BASE_DIR / "output_cleaned.csv"
HEADER_MAPPING_XLSX = BASE_DIR / "output_header_mapping.xlsx"


KEY_COLUMNS = ["dimension_1", "dimension_2", "dimension_3", "dimension_4"]
OUTPUT_COLUMNS = [
    "dimension_1",
    "dimension_2",
    "dimension_3",
    "dimension_4",
    "subject",
    "period",
    "category",
    "metric",
    "value",
    "source_sheet",
    "source_col_index",
]
HEADER_MAPPING_COLUMNS = [
    "source_sheet",
    "source_col_index",
    "source_header_row_1",
    "source_header_row_2",
    "source_header_row_3",
    "source_header_row_4",
    "subject",
    "period",
    "category",
    "metric",
]


SUBJECT_NORMALIZATION = {
    # Example:
    # "FY2526 1H DS": "DS",
}

METRIC_NORMALIZATION = {
    # Example:
    # "Other funding BU  include": "Other funding BU include",
}


def clean_text(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def forward_fill(values):
    filled = []
    current = ""
    for value in values:
        text = clean_text(value)
        if text:
            current = text
        filled.append(current)
    return filled


def normalize_subject(value):
    text = clean_text(value)
    return SUBJECT_NORMALIZATION.get(text, text)


def normalize_metric(value):
    text = clean_text(value)
    return METRIC_NORMALIZATION.get(text, text)


def build_derived_fields(output_row):
    """Customize this hook to add fields such as dates, shipment, or forecast values."""
    return output_row


def write_xlsx(path, headers, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "data"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


def write_csv(path, headers, rows):
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def transform_workbook():
    workbook = load_workbook(INPUT_FILE, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row_1 = forward_fill([worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)])
    header_row_2 = forward_fill([worksheet.cell(2, col).value for col in range(1, worksheet.max_column + 1)])
    header_row_3 = forward_fill([worksheet.cell(3, col).value for col in range(1, worksheet.max_column + 1)])
    header_row_4 = [clean_text(worksheet.cell(4, col).value) for col in range(1, worksheet.max_column + 1)]

    output_rows = []
    mapping_rows = []

    for col in range(5, worksheet.max_column + 1):
        subject = normalize_subject(header_row_1[col - 1])
        period = clean_text(header_row_2[col - 1])
        category = clean_text(header_row_3[col - 1])
        metric = normalize_metric(header_row_4[col - 1])

        mapping_rows.append([
            worksheet.title,
            col,
            clean_text(worksheet.cell(1, col).value),
            clean_text(worksheet.cell(2, col).value),
            clean_text(worksheet.cell(3, col).value),
            clean_text(worksheet.cell(4, col).value),
            subject,
            period,
            category,
            metric,
        ])

        for row in range(5, worksheet.max_row + 1):
            key_values = [clean_text(worksheet.cell(row, key_col).value) for key_col in range(1, 5)]
            if any(value == "" for value in key_values):
                continue

            value = worksheet.cell(row, col).value
            if value in (None, ""):
                continue

            output_row = [
                key_values[0],
                key_values[1],
                key_values[2],
                key_values[3],
                subject,
                period,
                category,
                metric,
                value,
                worksheet.title,
                col,
            ]
            output_rows.append(build_derived_fields(output_row))

    write_xlsx(OUTPUT_XLSX, OUTPUT_COLUMNS, output_rows)
    write_csv(OUTPUT_CSV, OUTPUT_COLUMNS, output_rows)
    write_xlsx(HEADER_MAPPING_XLSX, HEADER_MAPPING_COLUMNS, mapping_rows)

    print(f"Created {OUTPUT_XLSX.name} with {len(output_rows)} rows")
    print(f"Created {HEADER_MAPPING_XLSX.name} with {len(mapping_rows)} rows")


if __name__ == "__main__":
    transform_workbook()